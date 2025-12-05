from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models.distribucion_model import DistribucionResponse, AsignacionMaquina
import io


def generar_excel_estilo_maquina(asignacion: AsignacionMaquina, package_nombre: str, demanda: int) -> bytes:
    """
    Genera un archivo Excel con el estilo de UNA máquina específica
    Optimizado para que el técnico programe la máquina
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"ESTILO {asignacion.machine_nombre}"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11)
    
    # ==== SECCIÓN 1: INFORMACIÓN GENERAL ====
    ws['A1'] = f"ESTILO DE MÁQUINA: {asignacion.machine_nombre}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Tipo: {asignacion.tipo_maquina}"
    ws['A2'].font = subtitle_font
    ws['A3'] = f"Package: {package_nombre}"
    ws['A4'] = f"Demanda: {demanda} unidades"
    ws['A5'] = f"Tiempo usado: {asignacion.tiempo_total_usado}h de {asignacion.tiempo_disponible}h disponibles"
    ws['A6'] = f"Tiempo sobrante: {asignacion.tiempo_sobrante}h"
    
    # ==== SECCIÓN 2: PARTS ASIGNADOS ====
    ws['A8'] = "PARTS ASIGNADOS A ESTA MÁQUINA:"
    ws['A8'].font = subtitle_font
    
    row = 9
    for i, part in enumerate(asignacion.parts_asignados, 1):
        ws[f'A{row}'] = f"{i}. {part.part_number}"
        ws[f'C{row}'] = f"Cantidad: {part.cantidad_asignada}"
        ws[f'E{row}'] = f"Horas: {part.horas_corrida}h"
        row += 1
    
    # ==== SECCIÓN 3: TABLA DE ESTILO (PRINCIPAL) ====
    row += 1
    ws[f'A{row}'] = "CONFIGURACIÓN DE ESTACIONES Y HERRAMIENTAS"
    ws[f'A{row}'].font = title_font
    ws.merge_cells(f'A{row}:H{row}')
    
    row += 1
    headers = ["ESTACIÓN", "TIPO", "TOOL NUMBER", "ÁNGULO", "TIENE GUÍA", "AUTOINDEX", "PARTS QUE USAN"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    for estacion in asignacion.estilo:
        ws[f'A{row}'] = estacion.estacion
        ws[f'B{row}'] = estacion.tipo
        ws[f'C{row}'] = estacion.tool_number
        ws[f'D{row}'] = f"{estacion.angulo}°"
        ws[f'E{row}'] = "SÍ" if estacion.tiene_guia else "NO"
        ws[f'F{row}'] = "SÍ" if estacion.es_autoindex else "NO"
        ws[f'G{row}'] = ", ".join(estacion.parts_que_usan)
        row += 1
    
    # ==== SECCIÓN 4: HERRAMIENTAS FUERA DE ESTILO (SI HAY) ====
    if asignacion.estaciones_fuera_estilo:
        row += 1
        ws[f'A{row}'] = "⚠️ HERRAMIENTAS FUERA DEL ESTILO:"
        ws[f'A{row}'].font = Font(bold=True, color="C00000", size=12)
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        for estacion in asignacion.estaciones_fuera_estilo:
            ws[f'A{row}'] = estacion.estacion
            ws[f'B{row}'] = estacion.tipo
            ws[f'C{row}'] = estacion.tool_number
            ws[f'D{row}'] = f"{estacion.angulo}°"
            ws[f'E{row}'] = "SÍ" if estacion.tiene_guia else "NO"
            ws[f'F{row}'] = "SÍ" if estacion.es_autoindex else "NO"
            ws[f'G{row}'] = ", ".join(estacion.parts_que_usan)
            row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 60
    
    # Guardar en memoria
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()


def generar_excel_distribucion(distribucion: DistribucionResponse) -> bytes:
    """
    Genera un archivo Excel con los resultados de la distribución
    """
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # 1. HOJA: RESUMEN GENERAL
    crear_hoja_resumen(wb, distribucion)
    
    # 2. HOJA POR CADA MÁQUINA
    for asignacion in distribucion.asignaciones:
        crear_hoja_maquina(wb, asignacion, distribucion)
    
    # 3. HOJA: ESTILOS (todas las máquinas)
    crear_hoja_estilos(wb, distribucion)
    
    # 4. HOJA: ALERTAS Y ERRORES
    crear_hoja_alertas_errores(wb, distribucion)
    
    # Guardar en memoria
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()


def crear_hoja_resumen(wb: Workbook, dist: DistribucionResponse):
    """Crea hoja de resumen general"""
    ws = wb.create_sheet("RESUMEN GENERAL")
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # Título
    ws['A1'] = "REPORTE DE DISTRIBUCIÓN"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    # Información del Package
    ws['A3'] = "Package:"
    ws['B3'] = dist.package_nombre
    ws['A4'] = "Demanda:"
    ws['B4'] = dist.demanda
    ws['A5'] = "Horas Objetivo:"
    ws['B5'] = dist.horas_objetivo
    ws['A6'] = "Factible:"
    ws['B6'] = "SÍ" if dist.es_factible else "NO"
    ws['B6'].font = Font(bold=True, color="00B050" if dist.es_factible else "C00000")
    
    # Resumen
    ws['A8'] = "RESUMEN"
    ws['A8'].font = title_font
    ws['A9'] = "Total Máquinas:"
    ws['B9'] = dist.resumen.get("total_maquinas_usadas", 0)
    ws['A10'] = "Total Horas Productivas:"
    ws['B10'] = dist.resumen.get("total_horas_productivas", 0)
    ws['A11'] = "Eficiencia Promedio:"
    ws['B11'] = f"{dist.resumen.get('eficiencia_promedio', 0)}%"
    ws['A12'] = "Total Parts Distintos:"
    ws['B12'] = dist.resumen.get("total_parts_distintos", 0)
    
    # Tabla de máquinas
    ws['A14'] = "MÁQUINA"
    ws['B14'] = "TIPO"
    ws['C14'] = "PARTS ASIGNADOS"
    ws['D14'] = "HORAS USADAS"
    ws['E14'] = "HORAS DISPONIBLES"
    ws['F14'] = "EFICIENCIA %"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}14'].fill = header_fill
        ws[f'{col}14'].font = header_font
        ws[f'{col}14'].alignment = Alignment(horizontal='center')
    
    row = 15
    for asig in dist.asignaciones:
        ws[f'A{row}'] = asig.machine_nombre
        ws[f'B{row}'] = asig.tipo_maquina
        ws[f'C{row}'] = len(asig.parts_asignados)
        ws[f'D{row}'] = asig.tiempo_total_usado
        ws[f'E{row}'] = asig.tiempo_disponible
        eficiencia = (asig.tiempo_total_usado / asig.tiempo_disponible * 100) if asig.tiempo_disponible > 0 else 0
        ws[f'F{row}'] = f"{eficiencia:.1f}%"
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15


def crear_hoja_maquina(wb: Workbook, asignacion, dist: DistribucionResponse):
    """Crea hoja detallada para cada máquina"""
    ws = wb.create_sheet(f"{asignacion.machine_nombre}")
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=12)
    
    # Título
    ws['A1'] = f"MÁQUINA: {asignacion.machine_nombre}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Tipo: {asignacion.tipo_maquina}"
    ws['A3'] = f"Tiempo usado: {asignacion.tiempo_total_usado}h de {asignacion.tiempo_disponible}h"
    ws['A4'] = f"Tiempo sobrante: {asignacion.tiempo_sobrante}h"
    
    # Tabla de parts asignados
    ws['A6'] = "PART NUMBER"
    ws['B6'] = "CANTIDAD REQUERIDA"
    ws['C6'] = "CANTIDAD ASIGNADA"
    ws['D6'] = "HORAS CORRIDA"
    ws['E6'] = "ESTACIONES USADAS"
    ws['F6'] = "ESTACIONES UNIFICADAS"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}6'].fill = header_fill
        ws[f'{col}6'].font = header_font
        ws[f'{col}6'].alignment = Alignment(horizontal='center')
    
    row = 7
    for part in asignacion.parts_asignados:
        ws[f'A{row}'] = part.part_number
        ws[f'B{row}'] = part.cantidad_requerida
        ws[f'C{row}'] = part.cantidad_asignada
        ws[f'D{row}'] = part.horas_corrida
        ws[f'E{row}'] = part.estaciones_usadas
        ws[f'F{row}'] = part.estaciones_unificadas
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25


def crear_hoja_estilos(wb: Workbook, dist: DistribucionResponse):
    """Crea hoja con los estilos de todas las máquinas"""
    ws = wb.create_sheet("ESTILOS")
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    row = 1
    
    for asignacion in dist.asignaciones:
        # Título de máquina
        ws[f'A{row}'] = f"MÁQUINA: {asignacion.machine_nombre}"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        # Headers
        ws[f'A{row}'] = "ESTACIÓN"
        ws[f'B{row}'] = "TIPO"
        ws[f'C{row}'] = "TOOL NUMBER"
        ws[f'D{row}'] = "ÁNGULO"
        ws[f'E{row}'] = "TIENE GUÍA"
        ws[f'F{row}'] = "AUTOINDEX"
        ws[f'G{row}'] = "PARTS QUE USAN"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        # Estilo
        for estacion in asignacion.estilo:
            ws[f'A{row}'] = estacion.estacion
            ws[f'B{row}'] = estacion.tipo
            ws[f'C{row}'] = estacion.tool_number
            ws[f'D{row}'] = f"{estacion.angulo}°"
            ws[f'E{row}'] = "SÍ" if estacion.tiene_guia else "NO"
            ws[f'F{row}'] = "SÍ" if estacion.es_autoindex else "NO"
            ws[f'G{row}'] = ", ".join(estacion.parts_que_usan)  # Todos los parts
            row += 1
        
        # Estaciones fuera de estilo
        if asignacion.estaciones_fuera_estilo:
            row += 1
            ws[f'A{row}'] = "HERRAMIENTAS FUERA DEL ESTILO:"
            ws[f'A{row}'].font = Font(bold=True, color="C00000")
            row += 1
            
            for estacion in asignacion.estaciones_fuera_estilo:
                ws[f'A{row}'] = estacion.estacion
                ws[f'B{row}'] = estacion.tipo
                ws[f'C{row}'] = estacion.tool_number
                ws[f'D{row}'] = f"{estacion.angulo}°"
                ws[f'E{row}'] = "SÍ" if estacion.tiene_guia else "NO"
                ws[f'F{row}'] = "SÍ" if estacion.es_autoindex else "NO"
                ws[f'G{row}'] = ", ".join(estacion.parts_que_usan)  # Todos los parts
                row += 1
        
        row += 2  # Espacio entre máquinas
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 40


def crear_hoja_alertas_errores(wb: Workbook, dist: DistribucionResponse):
    """Crea hoja con alertas y errores"""
    ws = wb.create_sheet("ALERTAS Y ERRORES")
    
    title_font = Font(bold=True, size=12)
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Errores generales
    ws['A1'] = "ERRORES CRÍTICOS"
    ws['A1'].font = title_font
    ws['A1'].fill = error_fill
    ws.merge_cells('A1:B1')
    
    row = 2
    if dist.errores_generales:
        for error in dist.errores_generales:
            ws[f'A{row}'] = error
            ws[f'A{row}'].fill = error_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
    else:
        ws['A2'] = "Sin errores"
        ws['A2'].font = Font(color="00B050")
        row = 3
    
    # Alertas generales
    row += 1
    ws[f'A{row}'] = "ALERTAS"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = warning_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    if dist.alertas_generales:
        for alerta in dist.alertas_generales:
            ws[f'A{row}'] = alerta
            ws[f'A{row}'].fill = warning_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
    else:
        ws[f'A{row}'] = "Sin alertas"
        ws[f'A{row}'].font = Font(color="00B050")
    
    # Ajustar ancho
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 20
